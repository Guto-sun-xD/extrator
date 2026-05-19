#!/usr/bin/env python3
"""
Extrator PDF - Digitalização inteligente de documentos financeiros brasileiros
Versão 1.0.0

Lê PDFs da pasta entrada/, usa o Gemini para extrair dados estruturados
via visão computacional, e gera uma planilha Excel formatada em saida/.
"""

import json
import os
import re
import shutil
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional

import google.generativeai as genai
import openpyxl
from dotenv import load_dotenv
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, ValidationError, field_validator
from rich import box
from rich.align import Align
from rich.console import Console
from rich.panel import Panel
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
)
from rich.table import Table
from rich.text import Text

# ============================================================
# CONSTANTES
# ============================================================

VERSAO = "1.0.0"
MODELO_GEMINI = "gemini-2.5-flash"

PASTA_ENTRADA = Path("entrada")
PASTA_PROCESSADOS = Path("processados")
PASTA_SAIDA = Path("saida")
PASTA_NAO_IDENTIFICADOS = PASTA_ENTRADA / "nao_identificados"

# Paleta de cores para a planilha Excel
COR_CABECALHO_BG = "1F3864"       # azul escuro
COR_ZEBRA_ESCURO = "DCE6F1"       # azul claro (linhas pares)
COR_VENCIDO_BG = "FFB3B3"         # vermelho pastel
COR_PROXIMO_BG = "FFFF99"         # amarelo pastel

# Nomes amigáveis por tipo de documento
TIPOS_AMIGAVEIS: Dict[str, str] = {
    "boleto_bancario": "Boleto Bancário",
    "conta_consumo": "Conta de Consumo",
    "nota_fiscal": "Nota Fiscal",
    "comprovante_pagamento": "Comprovante de Pagamento",
    "recibo": "Recibo",
    "outros": "Outros",
}

# ============================================================
# CONFIGURAÇÃO DO RICH
# ============================================================

console = Console()

# ============================================================
# MODELOS PYDANTIC
# ============================================================

TipoDocumento = Literal[
    "boleto_bancario",
    "conta_consumo",
    "nota_fiscal",
    "comprovante_pagamento",
    "recibo",
    "outros",
]


class DocumentoExtraido(BaseModel):
    """Dados estruturados extraídos de um documento financeiro pelo Gemini."""

    # Campos presentes em todos os tipos de documento
    tipo: TipoDocumento
    fornecedor: str = Field(description="Nome do fornecedor ou emissor do documento")
    valor_total: float = Field(description="Valor total em reais (número float)")
    data_emissao: Optional[date] = Field(None)
    data_vencimento: Optional[date] = Field(None)
    categoria_sugerida: str = Field(description="Categoria financeira sugerida")
    descricao_outros: Optional[str] = Field(None)

    # Campos específicos para boleto_bancario
    codigo_barras: Optional[str] = Field(None)
    numero_documento: Optional[str] = Field(None)
    beneficiario: Optional[str] = Field(None)

    # Campos específicos para conta_consumo
    tipo_servico: Optional[str] = Field(None)
    mes_referencia: Optional[str] = Field(None)
    consumo: Optional[str] = Field(None)
    numero_instalacao_ou_cliente: Optional[str] = Field(None)

    # Campos específicos para nota_fiscal
    numero_nf: Optional[str] = Field(None)
    cnpj_emissor: Optional[str] = Field(None)
    descricao_produtos_servicos: Optional[str] = Field(None)

    # Metadados preenchidos durante o processamento (não vêm do Gemini)
    arquivo_original: Optional[str] = Field(None)
    arquivo_processado: Optional[str] = Field(None)

    @field_validator("valor_total", mode="before")
    @classmethod
    def normalizar_valor(cls, v: Any) -> float:
        """Aceita valores em formato brasileiro (R$ 1.234,56) ou float puro."""
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            v = v.replace("R$", "").replace(" ", "").strip()
            # Detecta formato brasileiro: ponto = milhar, vírgula = decimal
            if "," in v and "." in v:
                v = v.replace(".", "").replace(",", ".")
            elif "," in v:
                v = v.replace(",", ".")
            try:
                return float(v)
            except ValueError:
                return 0.0
        return 0.0

    @field_validator("data_emissao", "data_vencimento", mode="before")
    @classmethod
    def normalizar_data(cls, v: Any) -> Optional[date]:
        """Converte strings de data em objeto date, tolerando vários formatos."""
        if v is None or v in ("", "null"):
            return None
        if isinstance(v, date):
            return v
        if isinstance(v, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(v, fmt).date()
                except ValueError:
                    continue
        return None

    @field_validator("categoria_sugerida", mode="before")
    @classmethod
    def normalizar_categoria(cls, v: Any) -> str:
        """Mapeia variações ortográficas para os valores canônicos de categoria."""
        if not isinstance(v, str):
            return "outros"
        mapa = {
            "alimentação": "alimentação", "alimentacao": "alimentação",
            "moradia": "moradia",
            "transporte": "transporte",
            "lazer": "lazer",
            "saúde": "saúde", "saude": "saúde",
            "educação": "educação", "educacao": "educação",
            "serviços": "serviços", "servicos": "serviços",
            "outros": "outros",
        }
        return mapa.get(v.lower().strip(), "outros")


class ResultadoProcessamento(BaseModel):
    """Resultado (sucesso ou falha) do processamento de um único PDF."""

    pdf_path: str
    sucesso: bool
    documento: Optional[DocumentoExtraido] = None
    erro: Optional[str] = None


# ============================================================
# PROMPT PARA O GEMINI
# ============================================================

PROMPT_EXTRACAO = """\
Você é um especialista em análise de documentos financeiros brasileiros.
Analise cuidadosamente este documento PDF e extraia as informações financeiras.

CLASSIFICAÇÃO DO TIPO:
- "boleto_bancario"       → boleto bancário de cobrança
- "conta_consumo"         → energia elétrica, água, gás, internet, telefone, streaming
- "nota_fiscal"           → NF-e, NFS-e, NFCe ou nota em papel
- "comprovante_pagamento" → comprovante de TED, PIX, transferência
- "recibo"                → recibo de pagamento ou quitação
- "outros"                → qualquer outro tipo; explique em descricao_outros

Retorne APENAS um objeto JSON válido com esta estrutura (sem comentários):
{
    "tipo": "tipo_do_documento",
    "fornecedor": "Nome completo do fornecedor/emissor",
    "valor_total": 0.00,
    "data_emissao": "AAAA-MM-DD",
    "data_vencimento": "AAAA-MM-DD",
    "categoria_sugerida": "categoria",
    "descricao_outros": null,
    "codigo_barras": null,
    "numero_documento": null,
    "beneficiario": null,
    "tipo_servico": null,
    "mes_referencia": null,
    "consumo": null,
    "numero_instalacao_ou_cliente": null,
    "numero_nf": null,
    "cnpj_emissor": null,
    "descricao_produtos_servicos": null
}

REGRAS:
- valor_total: número float (ex: 150.50 — não "R$ 150,50")
- datas: formato AAAA-MM-DD; use null se ausente ou ilegível
- categoria_sugerida deve ser UMA de:
    alimentação, moradia, transporte, lazer, saúde, educação, serviços, outros
- Para boleto_bancario: preencha codigo_barras, numero_documento, beneficiario
- Para conta_consumo: preencha tipo_servico, mes_referencia, consumo, numero_instalacao_ou_cliente
- Para nota_fiscal: preencha numero_nf, cnpj_emissor, descricao_produtos_servicos
- Campos não aplicáveis ao tipo detectado devem ser null
- Retorne APENAS o JSON, sem texto antes ou depois

GUIA DE CATEGORIAS:
  energia elétrica/água/gás/aluguel/condomínio → moradia
  internet/telefone fixo → serviços
  streaming/assinaturas de lazer → lazer
  supermercado/restaurante/delivery → alimentação
  combustível/ônibus/metrô → transporte
  médico/farmácia/plano de saúde → saúde
  escola/curso/livros → educação
"""

# ============================================================
# EXIBIÇÃO NO TERMINAL (RICH)
# ============================================================


def exibir_banner() -> None:
    """Exibe o banner inicial da ferramenta."""
    texto = Text()
    texto.append("  EXTRATOR PDF  ", style="bold white on blue")
    texto.append(f"\n  v{VERSAO}  •  Documentos Financeiros Brasileiros  ", style="dim white")

    console.print()
    console.print(Panel(Align.center(texto), border_style="blue", padding=(1, 4)))
    console.print()


def exibir_documento(doc: DocumentoExtraido, pdf_path: Path) -> None:
    """Exibe painel resumido com os dados extraídos de um documento."""
    valor_fmt = _fmt_brl(doc.valor_total)
    data_e = _fmt_data(doc.data_emissao)
    data_v = _fmt_data(doc.data_vencimento)

    t = Table(show_header=False, box=box.SIMPLE, padding=(0, 1))
    t.add_column("campo", style="dim", width=22)
    t.add_column("valor", style="white")

    t.add_row("Tipo", f"[cyan]{TIPOS_AMIGAVEIS.get(doc.tipo, doc.tipo)}[/cyan]")
    t.add_row("Fornecedor", f"[bold]{doc.fornecedor}[/bold]")
    t.add_row("Valor Total", f"[bold green]{valor_fmt}[/bold green]")
    t.add_row("Emissão", data_e)
    t.add_row("Vencimento", data_v)
    t.add_row("Categoria", f"[yellow]{doc.categoria_sugerida}[/yellow]")

    if doc.tipo == "boleto_bancario":
        if doc.beneficiario:
            t.add_row("Beneficiário", doc.beneficiario)
        if doc.numero_documento:
            t.add_row("Nº Documento", doc.numero_documento)
    elif doc.tipo == "conta_consumo":
        if doc.tipo_servico:
            t.add_row("Serviço", doc.tipo_servico)
        if doc.mes_referencia:
            t.add_row("Mês Ref.", doc.mes_referencia)
        if doc.consumo:
            t.add_row("Consumo", doc.consumo)
    elif doc.tipo == "nota_fiscal":
        if doc.numero_nf:
            t.add_row("Nº NF", doc.numero_nf)
        if doc.cnpj_emissor:
            t.add_row("CNPJ Emissor", doc.cnpj_emissor)

    console.print(Panel(t, title=f"[bold blue]{pdf_path.name}[/bold blue]", border_style="green", padding=(0, 1)))


def exibir_erro(pdf_path: Path, mensagem: str) -> None:
    """Exibe painel de erro para um PDF com problema."""
    console.print(Panel(
        f"[red]{mensagem}[/red]",
        title=f"[bold red]Erro — {pdf_path.name}[/bold red]",
        border_style="red",
        padding=(0, 1),
    ))


def exibir_resumo_final(resultados: List[ResultadoProcessamento], planilha: Optional[Path]) -> None:
    """Exibe tabela de estatísticas ao final do processamento."""
    sucessos = [r for r in resultados if r.sucesso and r.documento]
    falhas = [r for r in resultados if not r.sucesso]
    total_valor = sum(r.documento.valor_total for r in sucessos)  # type: ignore[union-attr]

    tipos_count: Dict[str, int] = {}
    for r in sucessos:
        t = r.documento.tipo  # type: ignore[union-attr]
        tipos_count[t] = tipos_count.get(t, 0) + 1

    tab = Table(title="Resumo do Processamento", box=box.ROUNDED, border_style="blue")
    tab.add_column("Estatística", style="dim")
    tab.add_column("Valor", style="bold")

    tab.add_row("Processados com sucesso", f"[green]{len(sucessos)}[/green]")
    tab.add_row("Com erro", f"[red]{len(falhas)}[/red]")
    tab.add_row("Valor total extraído", f"[bold green]{_fmt_brl(total_valor)}[/bold green]")

    for tipo, qtd in sorted(tipos_count.items(), key=lambda x: -x[1]):
        tab.add_row(TIPOS_AMIGAVEIS.get(tipo, tipo), str(qtd))

    console.print()
    console.print(tab)

    if planilha:
        console.print()
        console.print(Panel(
            f"[bold green]Planilha gerada:[/bold green] [blue]{planilha}[/blue]",
            border_style="green",
        ))

    console.print()


# ============================================================
# HELPERS DE FORMATAÇÃO
# ============================================================


def _fmt_brl(valor: float) -> str:
    """Formata float como moeda brasileira: R$ 1.234,56"""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_data(d: Optional[date]) -> str:
    """Formata date em DD/MM/AAAA ou retorna string vazia."""
    return d.strftime("%d/%m/%Y") if d else ""


# ============================================================
# CONFIGURAÇÃO DO GEMINI
# ============================================================


def configurar_gemini() -> genai.GenerativeModel:
    """Lê GEMINI_API_KEY do .env e instancia o modelo Gemini."""
    load_dotenv()

    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        console.print("[red]Erro:[/red] GEMINI_API_KEY não encontrada.")
        console.print("Crie um arquivo [bold].env[/bold] com: GEMINI_API_KEY=sua_chave_aqui")
        sys.exit(1)

    genai.configure(api_key=api_key)

    return genai.GenerativeModel(
        model_name=MODELO_GEMINI,
        generation_config=genai.GenerationConfig(
            response_mime_type="application/json",
            temperature=0.1,  # baixa temperatura → respostas mais determinísticas
        ),
    )


# ============================================================
# EXTRAÇÃO DE DADOS COM GEMINI
# ============================================================


def _limpar_json(texto: str) -> str:
    """Remove blocos de código markdown que o modelo às vezes inclui."""
    texto = re.sub(r"```(?:json)?\s*", "", texto)
    return re.sub(r"```\s*$", "", texto).strip()


def extrair_dados_pdf(
    pdf_path: Path,
    modelo: genai.GenerativeModel,
    max_tentativas: int = 3,
) -> DocumentoExtraido:
    """
    Envia o PDF ao Gemini Files API e extrai os dados estruturados.

    Args:
        pdf_path: Caminho para o arquivo PDF a processar.
        modelo: Instância configurada do GenerativeModel.
        max_tentativas: Limite de retentativas em caso de erro 429 (rate limit).

    Returns:
        DocumentoExtraido validado pelo Pydantic.

    Raises:
        Exception: Erros não recuperáveis (PDF corrompido, API inválida, etc.).
    """
    arquivo_gemini = None

    try:
        # Faz upload do PDF para a Files API do Google
        arquivo_gemini = genai.upload_file(path=str(pdf_path), mime_type="application/pdf")

        # Aguarda o arquivo ficar disponível (estado ACTIVE)
        while arquivo_gemini.state.name == "PROCESSING":
            time.sleep(1)
            arquivo_gemini = genai.get_file(arquivo_gemini.name)

        if arquivo_gemini.state.name == "FAILED":
            raise ValueError("Falha no processamento do PDF pelo Google Files API.")

        # Chama o modelo com retry exponencial para erro 429
        for tentativa in range(1, max_tentativas + 1):
            try:
                resposta = modelo.generate_content([arquivo_gemini, PROMPT_EXTRACAO])
                break
            except Exception as exc:
                msg = str(exc).lower()
                eh_rate_limit = any(k in msg for k in ("429", "quota", "resource_exhausted"))

                if eh_rate_limit and tentativa < max_tentativas:
                    espera = 60 * tentativa  # 60s, 120s
                    console.print(f"  [yellow]Rate limit — aguardando {espera}s (tentativa {tentativa}/{max_tentativas})...[/yellow]")
                    time.sleep(espera)
                else:
                    raise

        dados = json.loads(_limpar_json(resposta.text))
        doc = DocumentoExtraido(**dados)
        doc.arquivo_original = pdf_path.name
        return doc

    finally:
        # Remove o arquivo da Files API para não acumular lixo
        if arquivo_gemini:
            try:
                genai.delete_file(arquivo_gemini.name)
            except Exception:
                pass


# ============================================================
# MOVIMENTAÇÃO DE ARQUIVOS
# ============================================================


def gerar_nome_arquivo(doc: DocumentoExtraido) -> str:
    """
    Gera nome padronizado: AAAA-MM-DD_Fornecedor_R$Valor.pdf

    Usa data de emissão quando disponível; cai para a data atual.
    """
    data = doc.data_emissao or date.today()
    data_str = data.strftime("%Y-%m-%d")

    fornecedor = re.sub(r"[^\w\s-]", "", doc.fornecedor, flags=re.UNICODE)
    fornecedor = re.sub(r"\s+", "_", fornecedor.strip())[:40]

    # Vírgula no valor evita conflito com extensão do arquivo
    valor_str = f"R${doc.valor_total:.2f}".replace(".", ",")

    return f"{data_str}_{fornecedor}_{valor_str}.pdf"


def mover_para_processados(pdf_path: Path, doc: DocumentoExtraido) -> Path:
    """Renomeia e move o PDF para processados/; evita sobrescrita com sufixo numérico."""
    PASTA_PROCESSADOS.mkdir(exist_ok=True)
    destino = PASTA_PROCESSADOS / gerar_nome_arquivo(doc)

    contador = 1
    while destino.exists():
        destino = PASTA_PROCESSADOS / f"{destino.stem}_{contador}.pdf"
        contador += 1

    shutil.move(str(pdf_path), str(destino))
    doc.arquivo_processado = destino.name
    return destino


def mover_para_nao_identificados(pdf_path: Path) -> None:
    """Move PDF problemático para entrada/nao_identificados/."""
    PASTA_NAO_IDENTIFICADOS.mkdir(parents=True, exist_ok=True)
    destino = PASTA_NAO_IDENTIFICADOS / pdf_path.name
    shutil.move(str(pdf_path), str(destino))


# ============================================================
# HELPERS DE FORMATAÇÃO EXCEL
# ============================================================


def _cabecalho(ws: Any, row: int, n_cols: int) -> None:
    """Aplica estilo de cabeçalho (fundo azul escuro, texto branco negrito)."""
    fill = PatternFill(start_color=COR_CABECALHO_BG, end_color=COR_CABECALHO_BG, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align


def _zebra(ws: Any, row: int, n_cols: int, par: bool) -> None:
    """Aplica fundo alternado (zebra) nas linhas de dados."""
    fill = (
        PatternFill(start_color=COR_ZEBRA_ESCURO, end_color=COR_ZEBRA_ESCURO, fill_type="solid")
        if par
        else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    )
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def _larguras(ws: Any, larguras: List[int]) -> None:
    """Define largura de cada coluna em caracteres."""
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg


def _destacar_vencimento(ws: Any, row: int, col: int, d: Optional[date]) -> None:
    """Pinta célula de vencimento: vermelho se vencido, amarelo se nos próximos 30 dias."""
    if d is None:
        return
    hoje = date.today()
    if d < hoje:
        cor = COR_VENCIDO_BG
    elif d <= hoje + timedelta(days=30):
        cor = COR_PROXIMO_BG
    else:
        return
    ws.cell(row=row, column=col).fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")


# ============================================================
# ABAS DA PLANILHA EXCEL
# ============================================================


def _aba_todos(ws: Any, docs: List[DocumentoExtraido]) -> None:
    """Popula a aba 'Todos' com a tabela completa de documentos."""
    headers = [
        "Tipo", "Fornecedor", "Valor Total (R$)", "Data Emissão", "Data Vencimento",
        "Categoria", "Arquivo Original", "Arquivo Processado",
        "Cód. Barras", "Nº Documento", "Beneficiário",
        "Tipo Serviço", "Mês Referência", "Consumo", "Nº Instalação/Cliente",
        "Nº NF", "CNPJ Emissor", "Descrição Produtos/Serviços",
    ]
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _cabecalho(ws, 1, len(headers))

    for i, doc in enumerate(docs, 2):
        par = i % 2 == 0
        row_data = [
            TIPOS_AMIGAVEIS.get(doc.tipo, doc.tipo), doc.fornecedor,
            doc.valor_total,
            _fmt_data(doc.data_emissao), _fmt_data(doc.data_vencimento),
            doc.categoria_sugerida,
            doc.arquivo_original or "", doc.arquivo_processado or "",
            doc.codigo_barras or "", doc.numero_documento or "", doc.beneficiario or "",
            doc.tipo_servico or "", doc.mes_referencia or "",
            doc.consumo or "", doc.numero_instalacao_ou_cliente or "",
            doc.numero_nf or "", doc.cnpj_emissor or "",
            doc.descricao_produtos_servicos or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)
        _zebra(ws, i, len(headers), par)
        ws.cell(row=i, column=3).number_format = 'R$ #,##0.00'
        _destacar_vencimento(ws, i, 5, doc.data_vencimento)

    _larguras(ws, [20, 30, 15, 12, 12, 15, 32, 32, 46, 15, 26, 15, 13, 13, 22, 10, 18, 42])


def _aba_boletos(ws: Any, docs: List[DocumentoExtraido]) -> None:
    """Popula a aba 'Boletos' apenas com documentos do tipo boleto_bancario."""
    boletos = [d for d in docs if d.tipo == "boleto_bancario"]
    headers = [
        "Fornecedor / Beneficiário", "Valor Total (R$)", "Data Emissão",
        "Data Vencimento", "Código de Barras", "Nº Documento", "Categoria", "Arquivo",
    ]
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _cabecalho(ws, 1, len(headers))

    for i, doc in enumerate(boletos, 2):
        par = i % 2 == 0
        row_data = [
            doc.beneficiario or doc.fornecedor, doc.valor_total,
            _fmt_data(doc.data_emissao), _fmt_data(doc.data_vencimento),
            doc.codigo_barras or "", doc.numero_documento or "",
            doc.categoria_sugerida, doc.arquivo_original or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)
        _zebra(ws, i, len(headers), par)
        ws.cell(row=i, column=2).number_format = 'R$ #,##0.00'
        _destacar_vencimento(ws, i, 4, doc.data_vencimento)

    _larguras(ws, [36, 15, 12, 12, 52, 20, 15, 36])


def _aba_contas(ws: Any, docs: List[DocumentoExtraido]) -> None:
    """Popula a aba 'Contas' apenas com documentos do tipo conta_consumo."""
    contas = [d for d in docs if d.tipo == "conta_consumo"]
    headers = [
        "Fornecedor", "Tipo Serviço", "Mês Referência", "Valor Total (R$)",
        "Data Vencimento", "Consumo", "Nº Instalação/Cliente", "Categoria", "Arquivo",
    ]
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _cabecalho(ws, 1, len(headers))

    for i, doc in enumerate(contas, 2):
        par = i % 2 == 0
        row_data = [
            doc.fornecedor, doc.tipo_servico or "", doc.mes_referencia or "",
            doc.valor_total, _fmt_data(doc.data_vencimento),
            doc.consumo or "", doc.numero_instalacao_ou_cliente or "",
            doc.categoria_sugerida, doc.arquivo_original or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)
        _zebra(ws, i, len(headers), par)
        ws.cell(row=i, column=4).number_format = 'R$ #,##0.00'

    _larguras(ws, [30, 18, 15, 15, 14, 15, 22, 15, 36])


def _aba_notas_fiscais(ws: Any, docs: List[DocumentoExtraido]) -> None:
    """Popula a aba 'Notas Fiscais' apenas com documentos do tipo nota_fiscal."""
    nfs = [d for d in docs if d.tipo == "nota_fiscal"]
    headers = [
        "Fornecedor / Emissor", "CNPJ Emissor", "Nº NF", "Valor Total (R$)",
        "Data Emissão", "Categoria", "Produtos / Serviços", "Arquivo",
    ]
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _cabecalho(ws, 1, len(headers))

    for i, doc in enumerate(nfs, 2):
        par = i % 2 == 0
        row_data = [
            doc.fornecedor, doc.cnpj_emissor or "", doc.numero_nf or "",
            doc.valor_total, _fmt_data(doc.data_emissao),
            doc.categoria_sugerida, doc.descricao_produtos_servicos or "",
            doc.arquivo_original or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)
        _zebra(ws, i, len(headers), par)
        ws.cell(row=i, column=4).number_format = 'R$ #,##0.00'

    _larguras(ws, [36, 20, 12, 15, 12, 15, 52, 36])


def _aba_outros(ws: Any, docs: List[DocumentoExtraido]) -> None:
    """Popula a aba 'Outros' com comprovantes, recibos e documentos não classificados."""
    outros = [d for d in docs if d.tipo not in ("boleto_bancario", "conta_consumo", "nota_fiscal")]
    headers = [
        "Tipo", "Fornecedor", "Valor Total (R$)", "Data Emissão",
        "Data Vencimento", "Categoria", "Descrição", "Arquivo",
    ]
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _cabecalho(ws, 1, len(headers))

    for i, doc in enumerate(outros, 2):
        par = i % 2 == 0
        row_data = [
            TIPOS_AMIGAVEIS.get(doc.tipo, doc.tipo), doc.fornecedor,
            doc.valor_total,
            _fmt_data(doc.data_emissao), _fmt_data(doc.data_vencimento),
            doc.categoria_sugerida, doc.descricao_outros or "",
            doc.arquivo_original or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)
        _zebra(ws, i, len(headers), par)
        ws.cell(row=i, column=3).number_format = 'R$ #,##0.00'

    _larguras(ws, [22, 30, 15, 12, 12, 15, 46, 36])


def _aba_resumo(ws: Any, docs: List[DocumentoExtraido]) -> None:
    """Cria a aba de resumo: estatísticas, gráfico de pizza e gráfico de barras."""
    hoje = date.today()

    # --- Título ---
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Resumo — Extração de Documentos Financeiros"
    ws["A1"].font = Font(bold=True, size=16, color=COR_CABECALHO_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws["A2"].value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="666666")

    # --- Estatísticas gerais ---
    ws["A4"].value = "Estatísticas Gerais"
    ws["A4"].font = Font(bold=True, size=13, color=COR_CABECALHO_BG)

    total_valor = sum(d.valor_total for d in docs)
    valores = [d.valor_total for d in docs]

    stats = [
        ("Total de documentos", len(docs)),
        ("Valor total", _fmt_brl(total_valor)),
        ("Menor valor", _fmt_brl(min(valores, default=0))),
        ("Maior valor", _fmt_brl(max(valores, default=0))),
        ("Valor médio", _fmt_brl(total_valor / len(docs) if docs else 0)),
    ]
    for i, (label, val) in enumerate(stats, 5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)

    # --- Distribuição por tipo (dados + gráfico de pizza) ---
    R_TIPO = 12  # linha de início desta seção
    ws.cell(row=R_TIPO, column=1, value="Distribuição por Tipo").font = Font(bold=True, size=13, color=COR_CABECALHO_BG)
    for col, h in enumerate(["Tipo", "Quantidade", "Valor Total (R$)"], 1):
        ws.cell(row=R_TIPO + 1, column=col, value=h).font = Font(bold=True)

    tipos_count: Dict[str, int] = {}
    tipos_valor: Dict[str, float] = {}
    for doc in docs:
        tipos_count[doc.tipo] = tipos_count.get(doc.tipo, 0) + 1
        tipos_valor[doc.tipo] = tipos_valor.get(doc.tipo, 0.0) + doc.valor_total

    r = R_TIPO + 2
    for tipo, qtd in sorted(tipos_count.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=TIPOS_AMIGAVEIS.get(tipo, tipo))
        ws.cell(row=r, column=2, value=qtd)
        ws.cell(row=r, column=3, value=tipos_valor.get(tipo, 0))
        ws.cell(row=r, column=3).number_format = 'R$ #,##0.00'
        r += 1

    if len(tipos_count) > 0:
        pizza = PieChart()
        pizza.title = "Documentos por Tipo"
        pizza.style = 10
        pizza.width = 15
        pizza.height = 10
        pizza.add_data(Reference(ws, min_col=2, min_row=R_TIPO + 1, max_row=r - 1), titles_from_data=True)
        pizza.set_categories(Reference(ws, min_col=1, min_row=R_TIPO + 2, max_row=r - 1))
        ws.add_chart(pizza, "E12")

    # --- Distribuição por categoria (dados + gráfico de barras) ---
    R_CAT = r + 2
    ws.cell(row=R_CAT, column=1, value="Distribuição por Categoria").font = Font(bold=True, size=13, color=COR_CABECALHO_BG)
    for col, h in enumerate(["Categoria", "Quantidade", "Valor Total (R$)"], 1):
        ws.cell(row=R_CAT + 1, column=col, value=h).font = Font(bold=True)

    cats_count: Dict[str, int] = {}
    cats_valor: Dict[str, float] = {}
    for doc in docs:
        cats_count[doc.categoria_sugerida] = cats_count.get(doc.categoria_sugerida, 0) + 1
        cats_valor[doc.categoria_sugerida] = cats_valor.get(doc.categoria_sugerida, 0.0) + doc.valor_total

    r2 = R_CAT + 2
    for cat, qtd in sorted(cats_count.items(), key=lambda x: -x[1]):
        ws.cell(row=r2, column=1, value=cat.capitalize())
        ws.cell(row=r2, column=2, value=qtd)
        ws.cell(row=r2, column=3, value=cats_valor.get(cat, 0))
        ws.cell(row=r2, column=3).number_format = 'R$ #,##0.00'
        r2 += 1

    if len(cats_count) > 0:
        barras = BarChart()
        barras.type = "col"
        barras.style = 10
        barras.title = "Valor por Categoria (R$)"
        barras.y_axis.title = "Valor (R$)"
        barras.x_axis.title = "Categoria"
        barras.width = 20
        barras.height = 12
        barras.add_data(Reference(ws, min_col=3, min_row=R_CAT + 1, max_row=r2 - 1), titles_from_data=True)
        barras.set_categories(Reference(ws, min_col=1, min_row=R_CAT + 2, max_row=r2 - 1))
        ws.add_chart(barras, f"E{R_CAT}")

    # --- Próximos vencimentos (30 dias) ---
    R_VENC = r2 + 2
    ws.cell(row=R_VENC, column=1, value="Próximos Vencimentos (30 dias)").font = Font(bold=True, size=13, color=COR_CABECALHO_BG)

    proximos = sorted(
        [d for d in docs if d.data_vencimento and hoje <= d.data_vencimento <= hoje + timedelta(days=30)],
        key=lambda d: d.data_vencimento,  # type: ignore[return-value]
    )

    if proximos:
        for col, h in enumerate(["Vencimento", "Fornecedor", "Tipo", "Valor (R$)"], 1):
            cell = ws.cell(row=R_VENC + 1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")

        for j, doc in enumerate(proximos, R_VENC + 2):
            dias = (doc.data_vencimento - hoje).days  # type: ignore[operator]
            cor = COR_VENCIDO_BG if dias <= 7 else COR_PROXIMO_BG
            for col, val in enumerate([
                _fmt_data(doc.data_vencimento), doc.fornecedor,
                TIPOS_AMIGAVEIS.get(doc.tipo, doc.tipo), doc.valor_total,
            ], 1):
                cell = ws.cell(row=j, column=col, value=val)
                cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            ws.cell(row=j, column=4).number_format = 'R$ #,##0.00'
    else:
        ws.cell(row=R_VENC + 1, column=1, value="Nenhum vencimento nos próximos 30 dias.")

    _larguras(ws, [36, 16, 18, 13, 18])


# ============================================================
# GERAÇÃO DA PLANILHA
# ============================================================


def gerar_planilha(docs: List[DocumentoExtraido]) -> Path:
    """
    Gera o arquivo Excel com todas as abas formatadas.

    Args:
        docs: Lista de documentos validados a incluir na planilha.

    Returns:
        Path do arquivo .xlsx gerado em saida/.
    """
    PASTA_SAIDA.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m_%H%M%S")
    caminho = PASTA_SAIDA / f"extrato_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrão "Sheet"

    ws_resumo = wb.create_sheet("Resumo")
    ws_todos = wb.create_sheet("Todos")
    ws_boletos = wb.create_sheet("Boletos")
    ws_contas = wb.create_sheet("Contas")
    ws_nfs = wb.create_sheet("Notas Fiscais")
    ws_outros = wb.create_sheet("Outros")

    _aba_resumo(ws_resumo, docs)
    _aba_todos(ws_todos, docs)
    _aba_boletos(ws_boletos, docs)
    _aba_contas(ws_contas, docs)
    _aba_notas_fiscais(ws_nfs, docs)
    _aba_outros(ws_outros, docs)

    wb.active = ws_resumo
    wb.save(str(caminho))
    return caminho


# ============================================================
# MAIN
# ============================================================


def main() -> None:
    """Orquestra: listar PDFs → extrair com Gemini → mover arquivos → gerar Excel."""
    exibir_banner()

    # Garante estrutura de pastas
    for pasta in (PASTA_ENTRADA, PASTA_PROCESSADOS, PASTA_SAIDA):
        pasta.mkdir(exist_ok=True)

    pdfs = sorted(PASTA_ENTRADA.glob("*.pdf"))

    if not pdfs:
        console.print(Panel(
            "[yellow]Nenhum PDF encontrado em [bold]entrada/[/bold]\n"
            "Coloque os documentos lá e execute novamente.[/yellow]",
            border_style="yellow",
        ))
        return

    # Exibe lista de PDFs encontrados
    tab = Table(title=f"{len(pdfs)} PDF(s) encontrado(s)", box=box.ROUNDED, border_style="blue")
    tab.add_column("#", style="dim", width=4)
    tab.add_column("Arquivo", style="white")
    tab.add_column("Tamanho", style="dim", justify="right")

    for i, pdf in enumerate(pdfs, 1):
        sz = pdf.stat().st_size
        sz_fmt = f"{sz / 1024:.1f} KB" if sz < 1_048_576 else f"{sz / 1_048_576:.1f} MB"
        tab.add_row(str(i), pdf.name, sz_fmt)

    console.print(tab)
    console.print()

    modelo = configurar_gemini()

    resultados: List[ResultadoProcessamento] = []
    documentos: List[DocumentoExtraido] = []

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        tarefa = progress.add_task("[blue]Processando PDFs...", total=len(pdfs))

        for pdf_path in pdfs:
            progress.update(tarefa, description=f"[blue]Processando: {pdf_path.name[:45]}...")

            try:
                doc = extrair_dados_pdf(pdf_path, modelo)
                console.print()
                exibir_documento(doc, pdf_path)
                mover_para_processados(pdf_path, doc)
                documentos.append(doc)
                resultados.append(ResultadoProcessamento(pdf_path=str(pdf_path), sucesso=True, documento=doc))

            except ValidationError as exc:
                msg = f"Dados inválidos — {exc.error_count()} erro(s) de validação Pydantic"
                console.print()
                exibir_erro(pdf_path, msg)
                resultados.append(ResultadoProcessamento(pdf_path=str(pdf_path), sucesso=False, erro=msg))
                mover_para_nao_identificados(pdf_path)

            except json.JSONDecodeError as exc:
                msg = f"Resposta do Gemini não é JSON válido: {exc}"
                console.print()
                exibir_erro(pdf_path, msg)
                resultados.append(ResultadoProcessamento(pdf_path=str(pdf_path), sucesso=False, erro=msg))
                mover_para_nao_identificados(pdf_path)

            except Exception as exc:
                msg = str(exc)
                console.print()
                exibir_erro(pdf_path, msg)
                resultados.append(ResultadoProcessamento(pdf_path=str(pdf_path), sucesso=False, erro=msg))
                # Não move para nao_identificados se o erro for de API (429, quota)
                if not any(k in msg.lower() for k in ("429", "quota", "resource_exhausted")):
                    mover_para_nao_identificados(pdf_path)

            finally:
                progress.advance(tarefa)

    planilha = None
    if documentos:
        console.print()
        with console.status("[blue]Gerando planilha Excel..."):
            planilha = gerar_planilha(documentos)

    exibir_resumo_final(resultados, planilha)


if __name__ == "__main__":
    main()
